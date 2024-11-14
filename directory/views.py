from django.db import IntegrityError
from rest_framework import viewsets
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from openpyxl import load_workbook
from .models import Material, Category
from .serializers import MaterialSerializer, CategorySerializer, CategoryWithTotalCostSerializer


# Вьюха для материалов
class MaterialViewSet(viewsets.ModelViewSet):
    queryset = Material.objects.all()
    serializer_class = MaterialSerializer


# Вьюха для категорий
class CategoryViewSet(viewsets.ModelViewSet):
    queryset = Category.objects.all()
    serializer_class = CategorySerializer


# Вьюха для загрузки материалов из Excel файла
class UploadMaterialsView(APIView):
    def post(self, request):
        file = request.FILES.get('file')
        if not file:
            return Response({"error": "No file provided"}, status=status.HTTP_400_BAD_REQUEST)

        # Загружаем Excel файл
        workbook = load_workbook(file)
        sheet = workbook.active

        # Проверяем, что файл содержит нужные заголовки
        headers = [cell.value for cell in sheet[1]]  # Чтение первой строки как заголовков
        required_headers = ['Материал', 'Категория', 'Код материала', 'Стоимость материала']

        if not all(header in headers for header in required_headers):
            return Response({"error": "Invalid Excel file format, missing required headers."},
                            status=status.HTTP_400_BAD_REQUEST)

        # Индексы столбцов
        material_index = headers.index('Материал')
        category_index = headers.index('Категория')
        code_index = headers.index('Код материала')
        cost_index = headers.index('Стоимость материала')

        # Чтение данных из Excel файла, начиная со второй строки (минус первая строка с заголовками)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            name = row[material_index]
            category_name = row[category_index]
            code = row[code_index]
            cost = row[cost_index]

            if not name or not category_name or not code or cost is None:
                continue  # Пропускаем строки с отсутствующими обязательными данными

            # Создание категории и материала
            try:
                category, created = Category.objects.get_or_create(name=category_name)
            except IntegrityError:
                pass
            try:
                Material.objects.create(name=name, category=category, code=code, cost=cost)
            except IntegrityError:
                pass

        return Response({"message": "Materials uploaded successfully"}, status=status.HTTP_201_CREATED)


class CategoryWithCostViewSet(viewsets.ModelViewSet):
    queryset = Category.objects.filter(parent=None)  # Начинаем с корневых категорий
    serializer_class = CategoryWithTotalCostSerializer
